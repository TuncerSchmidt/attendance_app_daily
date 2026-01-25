import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

# from procare_processor import process_procare
# from dhs_processor import process_dhs

from app.procare_processor import process_procare
from app.dhs_processor import process_dhs

# ================== COLORS ==================
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

COLOR_MAP = {
    "Swiped": GREEN,
    "Not Swiped": RED,
    "Void Transaction": YELLOW,
    "Void & Update Transaction": YELLOW,
    "Inform Parent": YELLOW,
    "Update Procare": YELLOW,
    "Not Swiped IN": RED,
    "Not Swiped OUT": RED,
    "Not Swiped BOTH": RED
}

# ================== TIME HELPERS ==================
def not_swiped_reason(p_in, p_out):
    if not p_in and p_out:
        return "Not Swiped IN"
    if p_in and not p_out:
        return "Not Swiped OUT"
    return "Not Swiped BOTH"
def parse_time(t):
    try:
        return datetime.strptime(t, "%H:%M").time()
    except:
        return None

def in_range(t, start, end):
    t = parse_time(t)
    return bool(t) and start <= t <= end

# ================== RESPONSE HELPERS ==================
def is_sa(resp): return bool(resp) and "(00) S/A" in resp
def is_b4(resp): return bool(resp) and "(B4)" in resp
def is_dd(resp): return bool(resp) and "(DD)" in resp
def is_card_not_active(resp): return bool(resp) and "Card Not Active" in resp

# ================== DHS TIME PICKER ==================
def pick_time_dhs(df, time_col, resp_col, pick):

    df = df[~df[resp_col].apply(is_card_not_active)]

    sa = df[df[resp_col].apply(is_sa)]
    source = sa if not sa.empty else None

    if source is None:
        b4 = df[df[resp_col].apply(is_b4)]
        source = b4 if not b4.empty else None

    if source is None or source.empty:
        return ""

    times = [parse_time(t) for t in source[time_col] if t]
    times = [t for t in times if t]

    if not times:
        return ""

    chosen = min(times) if pick == "min" else max(times)
    return chosen.strftime("%H:%M")

# ================== SLOT WINDOWS ==================
MORNING_START = parse_time("06:00")
MORNING_END   = parse_time("07:50")
AFTER_START   = parse_time("15:00")
AFTER_END     = parse_time("18:30")

# ==================================================
# ðŸ”¥ MAIN ORCHESTRATION FUNCTION
# ==================================================
def run_pipeline(
    procare_file,
    dhs_file,
    output_file
):
    # ---------- READ EXCELS (SADECE BURADA) ----------
    procare_top_rows = pd.read_excel(procare_file, header=None, nrows=3)
    procare_header = pd.read_excel(procare_file, header=None).iloc[0, 0]
    df_procare_raw = pd.read_excel(procare_file, header=8)

    df_dhs_raw = pd.read_excel(dhs_file, dtype=str)

    # ---------- PROCESS ----------
    procare = process_procare(df_procare_raw, procare_header).fillna("")
    dhs_raw = process_dhs(df_dhs_raw).fillna("")

    # ---------- NORMALIZE ----------
    procare["StudentID"] = procare["StudentID"].astype(str).str.strip()
    procare["Attdate"] = procare["Attdate"].astype(str).str.strip()

    dhs_raw["StudentID"] = dhs_raw["StudentID"].astype(str).str.strip()
    dhs_raw["Date"] = dhs_raw["Date"].astype(str).str.strip()

    # ---------- NORMALIZE DHS ----------
    dhs_rows = []

    for (sid, date), g in dhs_raw.groupby(["StudentID", "Date"]):
        dhs_rows.append({
            "StudentID": sid,
            "Date": date,
            "FullName": g.iloc[0]["FullName"],

            "Morning_IN": pick_time_dhs(g, "Morning_IN", "Morning_IN_Response", "min"),
            "Morning_OUT": pick_time_dhs(g, "Morning_OUT", "Morning_OUT_Response", "max"),

            "Afternoon_IN": pick_time_dhs(g, "Afternoon_IN", "Afternoon_IN_Response", "min"),
            "Afternoon_OUT": pick_time_dhs(g, "Afternoon_OUT", "Afternoon_OUT_Response", "max"),

            "Morning_IN_Response": " | ".join(g["Morning_IN_Response"].unique()),
            "Morning_OUT_Response": " | ".join(g["Morning_OUT_Response"].unique()),
            "Afternoon_IN_Response": " | ".join(g["Afternoon_IN_Response"].unique()),
            "Afternoon_OUT_Response": " | ".join(g["Afternoon_OUT_Response"].unique()),
        })

    dhs = pd.DataFrame(dhs_rows)

    # ---------- SLOT LOGIC ----------
    def process_slot(p_row, d_row, slot, start, end):
        p_in = p_row.get(f"{slot}_IN", "")
        p_out = p_row.get(f"{slot}_OUT", "")

        d_in = d_out = ""
        if d_row is not None:
            d_in = d_row.get(f"{slot}_IN", "")
            d_out = d_row.get(f"{slot}_OUT", "")

        final_in = p_in if p_in else d_in
        final_out = p_out if p_out else d_out

        has_procare_any = bool(p_in or p_out)
        has_procare_complete = bool(p_in and p_out)
        has_dhs_any = bool(d_in or d_out)
        has_dhs_complete = bool(d_in and d_out)

        if not has_procare_any:
            if has_dhs_any:
                return "Void Transaction", YELLOW, final_in, final_out
            return "", None, "", ""

        if p_in and not p_out and has_dhs_complete:
            return "Update Procare", YELLOW, final_in, final_out

        # if has_procare_any and not has_procare_complete:
        #     if has_dhs_any:
        #         return "Void Transaction", YELLOW, final_in, final_out
        #     return "Not Swiped", RED, final_in, final_out

        if has_procare_any and not has_procare_complete:
            if has_dhs_any:
                return "Void Transaction", YELLOW, final_in, final_out

            return not_swiped_reason(d_in, d_out), RED, final_in, final_out


        responses = []
        if d_row is not None:
            responses = [
                d_row.get(f"{slot}_IN_Response", ""),
                d_row.get(f"{slot}_OUT_Response", "")
            ]

        if all(is_dd(r) for r in responses if r):
            return not_swiped_reason(d_in, d_out), RED, final_in, final_out

        if not has_dhs_complete:
            return not_swiped_reason(d_in, d_out), RED, final_in, final_out

        valid = in_range(p_in, start, end) and in_range(p_out, start, end)

        if valid:
            if any(is_b4(r) for r in responses):
                return "Inform Parent", YELLOW, final_in, final_out
            return "Swiped", GREEN, final_in, final_out

        return "Void & Update Transaction", YELLOW, final_in, final_out

    # ---------- MAIN LOOP ----------
    rows = []
    processed = set()

    for _, p in procare.iterrows():
        sid, date = p["StudentID"], p["Attdate"]
        d_match = dhs[(dhs["StudentID"] == sid) & (dhs["Date"] == date)]
        d_row = d_match.iloc[0] if not d_match.empty else None

        m = process_slot(p, d_row, "Morning", MORNING_START, MORNING_END)
        a = process_slot(p, d_row, "Afternoon", AFTER_START, AFTER_END)

        rows.append({
            "Full Name": p["Full Name"],
            "StudentID": sid,
            "Date": date,
            "Morning_IN": m[2],
            "Morning_OUT": m[3],
            "Morning_Response": m[0],
            "Afternoon_IN": a[2],
            "Afternoon_OUT": a[3],
            "Afternoon_Response": a[0],
            "M_Color": m[1],
            "A_Color": a[1]
        })

        processed.add((sid, date))

    # ---------- DHS ONLY ----------
    # ---------- DHS ONLY (ORÄ°JÄ°NAL DAVRANIÅž KORUNDU) ----------
    for _, d in dhs.iterrows():
        key = (d["StudentID"], d["Date"])
        if key in processed:
            continue

        has_morning = bool(d["Morning_IN"] or d["Morning_OUT"])
        has_afternoon = bool(d["Afternoon_IN"] or d["Afternoon_OUT"])

        rows.append({
            "Full Name": d["FullName"],
            "StudentID": d["StudentID"],
            "Date": d["Date"],

            "Morning_IN": d["Morning_IN"] if has_morning else "",
            "Morning_OUT": d["Morning_OUT"] if has_morning else "",
            "Morning_Response": "Void Transaction" if has_morning else "",

            "Afternoon_IN": d["Afternoon_IN"] if has_afternoon else "",
            "Afternoon_OUT": d["Afternoon_OUT"] if has_afternoon else "",
            "Afternoon_Response": "Void Transaction" if has_afternoon else "",

            "M_Color": YELLOW if has_morning else None,
            "A_Color": YELLOW if has_afternoon else None
        })


    # ---------- WRITE FINAL ----------
    df = pd.DataFrame(rows)
    df = df.sort_values(by="Full Name", kind="stable").reset_index(drop=True)

    df.drop(columns=["M_Color", "A_Color"]).to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    for i, row in df.iterrows():
        r = i + 2

        # if row["M_Color"]:
        #     ws[f"D{r}"].fill = row["M_Color"]
        #     ws[f"E{r}"].fill = row["M_Color"]

        # if row["Morning_Response"] in COLOR_MAP:
        #     ws[f"F{r}"].fill = COLOR_MAP[row["Morning_Response"]]

        # if row["A_Color"]:
        #     ws[f"G{r}"].fill = row["A_Color"]
        #     ws[f"H{r}"].fill = row["A_Color"]

        # if row["Afternoon_Response"] in COLOR_MAP:
        #     ws[f"I{r}"].fill = COLOR_MAP[row["Afternoon_Response"]]

        # ===== MORNING =====
        mr = row["Morning_Response"]

        if mr == "Not Swiped IN":
            ws[f"D{r}"].fill = RED      # IN
            ws[f"E{r}"].fill = GREEN    # OUT

        elif mr == "Not Swiped OUT":
            ws[f"D{r}"].fill = GREEN
            ws[f"E{r}"].fill = RED

        elif mr == "Not Swiped BOTH":
            ws[f"D{r}"].fill = RED
            ws[f"E{r}"].fill = RED

        else:
            if row["M_Color"]:
                ws[f"D{r}"].fill = row["M_Color"]
                ws[f"E{r}"].fill = row["M_Color"]

        if mr in COLOR_MAP:
            ws[f"F{r}"].fill = COLOR_MAP[mr]

        # ===== AFTERNOON =====
        ar = row["Afternoon_Response"]

        if ar == "Not Swiped IN":
            ws[f"G{r}"].fill = RED
            ws[f"H{r}"].fill = GREEN

        elif ar == "Not Swiped OUT":
            ws[f"G{r}"].fill = GREEN
            ws[f"H{r}"].fill = RED

        elif ar == "Not Swiped BOTH":
            ws[f"G{r}"].fill = RED
            ws[f"H{r}"].fill = RED

        else:
            if row["A_Color"]:
                ws[f"G{r}"].fill = row["A_Color"]
                ws[f"H{r}"].fill = row["A_Color"]

        if ar in COLOR_MAP:
            ws[f"I{r}"].fill = COLOR_MAP[ar]

            
    ws.insert_rows(1, amount=3)

    bold_font = Font(bold=True)

    # procare ilk 3 satÄ±rÄ± yaz + bold
    for r in range(3):
        for c, val in enumerate(procare_top_rows.iloc[r]):
            cell = ws.cell(row=r + 1, column=c + 1, value=val)
            cell.font = bold_font
            
    wb.save(output_file)
